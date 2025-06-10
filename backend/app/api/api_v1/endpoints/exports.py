"""
Export functionality for reports and data
"""
from fastapi import APIRouter, HTTPException, Depends, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from typing import Optional, Dict, Any
from datetime import datetime, date, timedelta
import io
import csv
import json
import logging

# For PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# For Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.core.database import get_db
from app.models.database import Patient, Prediction, AuditLog
from app.models.schemas import ExportRequest, ExportResponse

router = APIRouter()
logger = logging.getLogger(__name__)

@router.post("/patients/csv", response_model=ExportResponse)
async def export_patients_csv(
    export_request: ExportRequest,
    db: Session = Depends(get_db)
):
    """Export patients data as CSV"""
    try:
        query = db.query(Patient)
        
        # Apply date filters
        if export_request.start_date:
            query = query.filter(Patient.created_at >= export_request.start_date)
        if export_request.end_date:
            query = query.filter(Patient.created_at <= export_request.end_date)
        
        patients = query.all()
        
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'ID', 'Patient ID', 'First Name', 'Last Name', 'Age', 'Gender',
            'Phone', 'Email', 'Medical Record Number', 'Created At'
        ]
        writer.writerow(headers)
        
        # Data rows
        for patient in patients:
            writer.writerow([
                patient.id,
                patient.patient_id,
                patient.first_name,
                patient.last_name,
                patient.age,
                patient.gender,
                patient.phone,
                patient.email,
                patient.medical_record_number,
                patient.created_at.isoformat() if patient.created_at else ""
            ])
        
        # Create response
        csv_data = output.getvalue()
        filename = f"patients_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            io.StringIO(csv_data),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting patients CSV: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.post("/predictions/csv")
async def export_predictions_csv(
    export_request: ExportRequest,
    db: Session = Depends(get_db)
):
    """Export predictions data as CSV"""
    try:
        query = db.query(Prediction).join(Patient, Prediction.patient_id == Patient.id, isouter=True)
        
        # Apply date filters
        if export_request.start_date:
            query = query.filter(Prediction.created_at >= export_request.start_date)
        if export_request.end_date:
            query = query.filter(Prediction.created_at <= export_request.end_date)
        
        predictions = query.all()
        
        # Generate CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'Prediction ID', 'Patient ID', 'Patient Name', 'Image Filename',
            'Prediction', 'Confidence', 'Inference Time', 'Clinical Notes',
            'Reviewed', 'Reviewed By', 'Created At'
        ]
        writer.writerow(headers)
        
        # Data rows
        for prediction in predictions:
            patient_name = ""
            if prediction.patient:
                patient_name = f"{prediction.patient.first_name} {prediction.patient.last_name}"
            
            writer.writerow([
                prediction.id,
                prediction.patient_id,
                patient_name,
                prediction.original_filename,
                prediction.prediction,
                prediction.confidence,
                prediction.inference_time,
                prediction.clinical_notes,
                prediction.reviewed,
                prediction.reviewed_by,
                prediction.created_at.isoformat() if prediction.created_at else ""
            ])
        
        # Create response
        csv_data = output.getvalue()
        filename = f"predictions_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            io.StringIO(csv_data),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting predictions CSV: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.post("/patients/excel")
async def export_patients_excel(
    export_request: ExportRequest,
    db: Session = Depends(get_db)
):
    """Export patients data as Excel file"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=501, detail="Excel export not available. Install openpyxl.")
    
    try:
        query = db.query(Patient)
        
        # Apply date filters
        if export_request.start_date:
            query = query.filter(Patient.created_at >= export_request.start_date)
        if export_request.end_date:
            query = query.filter(Patient.created_at <= export_request.end_date)
        
        patients = query.all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patients"
        
        # Headers
        headers = [
            'ID', 'Patient ID', 'First Name', 'Last Name', 'Age', 'Gender',
            'Phone', 'Email', 'Medical Record Number', 'Created At'
        ]
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, patient in enumerate(patients, 2):
            ws.cell(row=row, column=1, value=patient.id)
            ws.cell(row=row, column=2, value=patient.patient_id)
            ws.cell(row=row, column=3, value=patient.first_name)
            ws.cell(row=row, column=4, value=patient.last_name)
            ws.cell(row=row, column=5, value=patient.age)
            ws.cell(row=row, column=6, value=patient.gender)
            ws.cell(row=row, column=7, value=patient.phone)
            ws.cell(row=row, column=8, value=patient.email)
            ws.cell(row=row, column=9, value=patient.medical_record_number)
            ws.cell(row=row, column=10, value=patient.created_at.isoformat() if patient.created_at else "")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"patients_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting patients Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.post("/report/pdf")
async def generate_summary_report_pdf(
    export_request: ExportRequest,
    db: Session = Depends(get_db)
):
    """Generate comprehensive PDF report"""
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(status_code=501, detail="PDF export not available. Install reportlab.")
    
    try:
        # Get data
        end_date = export_request.end_date or datetime.now().date()
        start_date = export_request.start_date or (end_date - timedelta(days=30))
        
        # Query data
        patients = db.query(Patient).filter(
            Patient.created_at >= start_date,
            Patient.created_at <= end_date
        ).all()
        
        predictions = db.query(Prediction).filter(
            Prediction.created_at >= start_date,
            Prediction.created_at <= end_date
        ).all()
        
        # Calculate statistics
        total_patients = len(patients)
        total_predictions = len(predictions)
        pneumonia_cases = len([p for p in predictions if p.prediction == "PNEUMONIA"])
        normal_cases = len([p for p in predictions if p.prediction == "NORMAL"])
        avg_confidence = sum(p.confidence for p in predictions) / total_predictions if total_predictions > 0 else 0
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        story.append(Paragraph("Pneumonia Detection System Report", title_style))
        story.append(Spacer(1, 12))
        
        # Report period
        story.append(Paragraph(f"Report Period: {start_date} to {end_date}", styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Summary statistics
        story.append(Paragraph("Summary Statistics", styles['Heading2']))
        
        stats_data = [
            ['Metric', 'Value'],
            ['Total Patients', str(total_patients)],
            ['Total Predictions', str(total_predictions)],
            ['Pneumonia Cases', str(pneumonia_cases)],
            ['Normal Cases', str(normal_cases)],
            ['Average Confidence', f"{avg_confidence:.1%}"],
        ]
        
        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # Recent predictions (last 10)
        if predictions:
            story.append(Paragraph("Recent Predictions", styles['Heading2']))
            
            pred_data = [['Date', 'Patient', 'Prediction', 'Confidence']]
            recent_predictions = sorted(predictions, key=lambda x: x.created_at, reverse=True)[:10]
            
            for pred in recent_predictions:
                patient_name = "Unknown"
                if pred.patient:
                    patient_name = f"{pred.patient.first_name} {pred.patient.last_name}"
                
                pred_data.append([
                    pred.created_at.strftime('%Y-%m-%d'),
                    patient_name,
                    pred.prediction,
                    f"{pred.confidence:.1%}"
                ])
            
            pred_table = Table(pred_data)
            pred_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(pred_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        filename = f"pneumonia_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error generating PDF report: {e}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")

@router.get("/formats")
async def get_available_export_formats():
    """Get list of available export formats"""
    formats = {
        "csv": {
            "name": "CSV",
            "description": "Comma-separated values",
            "available": True
        },
        "excel": {
            "name": "Excel",
            "description": "Microsoft Excel format",
            "available": OPENPYXL_AVAILABLE
        },
        "pdf": {
            "name": "PDF",
            "description": "Portable Document Format",
            "available": REPORTLAB_AVAILABLE
        }
    }
    
    return {
        "formats": formats,
        "missing_dependencies": {
            "openpyxl": not OPENPYXL_AVAILABLE,
            "reportlab": not REPORTLAB_AVAILABLE
        }
    }
